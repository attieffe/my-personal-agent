#!/usr/bin/env python3
"""
pdf_ricevute_mediche.py
Estrae dati da un PDF di ricevute mediche scansionate e produce un file XLSX.

Strategia: Claude Sonnet Vision per ogni pagina.
Checkpoint JSON progressivo: se interrotto, riprende dall'ultima pagina salvata.

Uso:
    python pdf_ricevute_mediche.py [input.pdf] [output.xlsx]
    python pdf_ricevute_mediche.py input.pdf output.xlsx --reset   # riparte da zero

Dipendenze:
    pip install pymupdf openpyxl pillow anthropic
"""

import base64
import io
import json
import logging
import re
import sys
import time
from difflib import get_close_matches
from pathlib import Path

import fitz
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import anthropic


def _get_anthropic_client() -> anthropic.Anthropic:
    import os
    if os.environ.get("ANTHROPIC_API_KEY"):
        return anthropic.Anthropic()
    creds_path = Path.home() / ".claude" / ".credentials.json"
    if creds_path.exists():
        with open(creds_path) as f:
            tok = json.load(f).get("claudeAiOauth", {}).get("accessToken", "")
        if tok:
            return anthropic.Anthropic(auth_token=tok)
    raise RuntimeError("Nessuna credenziale Anthropic trovata.")


# ─── CONFIGURAZIONE ────────────────────────────────────────────────────────────

INPUT_PDF   = "ricevute.pdf"
OUTPUT_XLSX = "ricevute_mediche.xlsx"

KNOWN_CF = {
    "FMNTTL87T11D976U": "Attilio",
    "MSCCHR81M66E951K": "Chiara",
    "FMNLSN20E27D286P": "Alessandro",
    "FMNLCA22L65B729C": "Alice",
}

MODEL      = "claude-haiku-4-5-20251001"
DPI        = 300
DELAY_SEC  = 1.5   # pausa tra chiamate

# ─── LOGGING ───────────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ─── NORMALIZZAZIONE ───────────────────────────────────────────────────────────

def normalize_date(raw: str) -> str:
    if not raw:
        return ""
    raw = raw.strip()
    m = re.match(r"(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})", raw)
    if m:
        return f"{int(m.group(3)):02d}/{int(m.group(2)):02d}/{m.group(1)}"
    m = re.match(r"(\d{1,2})[-/\.\s](\d{1,2})[-/\.\s](\d{2,4})", raw)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        # Sanity check: giorno 1-31, mese 1-12, anno ragionevole
        if 1 <= int(d) <= 31 and 1 <= int(mo) <= 12 and 2000 <= int(y) <= 2030:
            return f"{int(d):02d}/{int(mo):02d}/{y}"
    return raw


def normalize_amount(raw: str) -> str:
    if not raw:
        return ""
    raw = raw.strip().replace("€", "").replace(" ", "")
    m = re.match(r"^(\d{1,3})\.(\d{2})$", raw)
    if m:
        return f"{m.group(1)},{m.group(2)}"
    return raw


def resolve_cf(raw_cf: str) -> tuple:
    if not raw_cf:
        return "", "?"
    raw_cf = re.sub(r"[^A-Z0-9]", "", raw_cf.upper())
    if not raw_cf:
        return "", "?"
    if raw_cf in KNOWN_CF:
        return raw_cf, KNOWN_CF[raw_cf]
    candidates = get_close_matches(raw_cf, KNOWN_CF.keys(), n=1, cutoff=0.72)
    if candidates:
        cf = candidates[0]
        log.info(f"          fuzzy: {raw_cf} → {cf} ({KNOWN_CF[cf]})")
        return cf, KNOWN_CF[cf]
    return raw_cf, "?"


# ─── VISION ────────────────────────────────────────────────────────────────────

PROMPT = """Sei un estrattore preciso di dati da ricevute mediche italiane.
Analizza questa immagine ed estrai i 4 campi con la massima accuratezza.

CAMPI:
1. **data** — data del documento (formato output: GG/MM/AAAA)
2. **cf** — codice fiscale del PAZIENTE/ASSISTITO (16 caratteri: es. RSSMRA80A01H501U). Cerca etichette: "C.F.", "Cod. Fisc.", "PAZIENTE", "ASSISTITO". NON il CF del medico o farmacista.
3. **importo** — totale pagato (formato output: cifre con VIRGOLA, es: 45,50)
4. **struttura** — nome della farmacia, studio o struttura sanitaria EMITTENTE

Rispondi SOLO con JSON:
{"data": "GG/MM/AAAA", "cf": "16CARATTERI", "importo": "X,XX", "struttura": "nome"}

Usa null per campi assenti o illeggibili."""


def image_to_b64(img: Image.Image) -> str:
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return base64.standard_b64encode(buf.getvalue()).decode()


def call_vision(client, img: Image.Image, page_num: int) -> dict:
    b64 = image_to_b64(img)
    for attempt in range(5):
        try:
            msg = client.messages.create(
                model=MODEL,
                max_tokens=300,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                        {"type": "text", "text": PROMPT},
                    ],
                }],
            )
            raw = msg.content[0].text.strip()
            m = re.search(r"\{.*?\}", raw, re.DOTALL)
            if m:
                return json.loads(m.group())
            return {}
        except Exception as e:
            if "429" in str(e) and attempt < 4:
                wait = 10 * (attempt + 1)
                log.warning(f"  p{page_num}: rate limit → attendo {wait}s (tentativo {attempt+1}/5)...")
                time.sleep(wait)
            else:
                log.error(f"  p{page_num}: errore: {e}")
                return {}
    return {}


# ─── PIPELINE ──────────────────────────────────────────────────────────────────

def process_page(client, page_num: int, img: Image.Image) -> dict:
    log.info(f"Pagina {page_num:>3}: Sonnet Vision (DPI {DPI})...")
    time.sleep(DELAY_SEC)

    llm = call_vision(client, img, page_num)

    cf_raw = (llm.get("cf") or "").strip()
    cf, persona = resolve_cf(cf_raw)

    return {
        "pagina":    page_num,
        "data":      normalize_date(llm.get("data") or ""),
        "cf":        cf,
        "persona":   persona,
        "importo":   normalize_amount(llm.get("importo") or ""),
        "struttura": (llm.get("struttura") or "").strip(),
    }


def write_xlsx(rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ricevute"

    headers = ["Pagina", "Data", "Codice Fiscale", "Persona", "Importo (€)", "Struttura/Farmacia"]
    hfill = PatternFill("solid", fgColor="2E74B5")
    hfont = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row["pagina"])
        ws.cell(r, 2, row["data"])
        ws.cell(r, 3, row["cf"])
        ws.cell(r, 4, row["persona"])
        ws.cell(r, 5, row["importo"])
        ws.cell(r, 6, row["struttura"])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    log.info(f"Salvato: {output_path}")


def main():
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else INPUT_PDF
    out_path = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_XLSX
    reset    = "--reset" in sys.argv

    checkpoint = Path(out_path).with_suffix(".checkpoint.json")

    if not Path(pdf_path).exists():
        log.error(f"File non trovato: {pdf_path}")
        sys.exit(1)

    client = _get_anthropic_client()

    # Carica checkpoint se esiste
    done: dict = {}
    if checkpoint.exists() and not reset:
        done = {r["pagina"]: r for r in json.loads(checkpoint.read_text())}
        log.info(f"Ripresa da checkpoint: {len(done)} pagine già processate")

    log.info(f"PDF: {pdf_path} | DPI: {DPI} | modello: {MODEL}")
    doc = fitz.open(pdf_path)
    total = len(doc)
    log.info(f"Pagine totali: {total}")

    for i, page in enumerate(doc, 1):
        if i in done:
            log.info(f"Pagina {i:>3}: già in checkpoint, salto")
            continue
        pix = page.get_pixmap(dpi=DPI)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        row = process_page(client, i, img)
        log.info(f"          → {row}")
        done[i] = row
        # Salva checkpoint dopo ogni pagina
        checkpoint.write_text(json.dumps(list(done.values()), ensure_ascii=False, indent=2))

    doc.close()

    rows = [done[i] for i in sorted(done.keys())]
    write_xlsx(rows, out_path)

    found_cf = sum(1 for r in rows if r["persona"] != "?")
    log.info(f"Completato. {len(rows)} ricevute | CF: {found_cf}/{len(rows)} ({round(found_cf/len(rows)*100)}%)")
    if checkpoint.exists():
        checkpoint.unlink()


if __name__ == "__main__":
    main()
